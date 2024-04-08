from openpyxl import load_workbook, Workbook
from openpyxl.styles import PatternFill


def entendendo_openpyxl():
    # Criando uma planilha:
    # planilha = Workbook()

    # Abre a planilha desejada:
    planilha = load_workbook('Planilhas/final.xlsx')
    print(planilha)

    # Pega a aba desejada da planilha:
    # aba_desejada = planilha['Aba2']
    # print(aba_desejada)

    # Criando uma aba na planilha:
    # nova_aba = planilha.create_sheet(title='Aba3')

    # Pega a aba da planilha que abre com ela:
    aba_ativa = planilha.active
    print(aba_ativa)

    lista_valores_altos = []
    validador = False
    for linha in aba_ativa['A1:D37']:  # -> Percorre as linhas presentes no intervalo informado.
        print(linha)
        print()
        for celula in linha:  # -> Percorre cada celula da linha atual.
            if celula.column == 2:  # 1 = A | 2 = B | 3 = C
                # celula.value = 'Uhul!'

                if ',' in celula.value:  # -> Verifico se existe uma virgula na celula atual.
                    celula.value = celula.value.replace(',', '.')  # -> Se existir uma virgula na celula eu troco para ponto (.)

                if float(celula.value) > 20:
                    validador = True
                    celula.fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')  # -> eb4034
            print(celula.value)

        if validador:
            lista_valores_altos.append(elemento.value for elemento in linha)
            validador = False

        print()

    nova_aba = planilha['Aba2']  # -> Trocando para a aba "Aba2".
    for elemento in lista_valores_altos:
        nova_aba.append(elemento)

    planilha.save('Planilhas/Outra.xlsx')  # -> Salvando a planilha alterada no caminho informado.
    planilha.close()  # -> Fechando a planilha após fazer tudo com ela.
